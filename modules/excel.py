import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, Protection
from openpyxl.utils import get_column_letter

from modules.helpers import cpl_pis_of, grade_of_score, so_pi_labels


def _compute_total(scores_map, components):
    s = 0
    for comp in components:
        val = scores_map.get(comp["id"])
        if val is not None:
            s += val * (comp["weight"] / 100.0)
    return s


def _grade_formula(cell):
    return (
        f'IF({cell}="","",IF({cell}>=85,"A",IF({cell}>=75,"B",'
        f'IF({cell}>=65,"C",IF({cell}>=50,"D","E")))))'
    )


def build_workbook(data, include_formulas=True):
    components = data["components"]
    categories = data["categories"]
    students = data["students"]
    score_rows = {}
    for s in data["scores"]:
        score_rows.setdefault(s["student_id"], {})[s["component_id"]] = s["score"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Penilaian"
    thin = Side(style="thin", color="cbd5e1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="eef2f8")
    group_fill = PatternFill("solid", fgColor="dbe4f0")
    head_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    first_comp_col = 4
    last_comp_col = first_comp_col + len(components) - 1
    total_col = last_comp_col + 1
    grade_col = total_col + 1
    last_col = grade_col

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    tcell = ws.cell(row=1, column=1, value=(
        f"{data['course']['course_code']} · {data['course']['course_name']} — "
        f"Kelas {data['class']['name']} | {data['class']['lecturer']} | "
        f"SKS {data['course']['sks']} | {data['course']['semester']}"
    ))
    tcell.font = Font(bold=True, size=12)
    tcell.alignment = center

    ws.cell(row=2, column=1, value="No")
    ws.cell(row=2, column=2, value="NIM")
    ws.cell(row=2, column=3, value="Nama")
    col = first_comp_col
    for cat in categories:
        span = len(cat["components"])
        if span > 1:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + span - 1)
        cell = ws.cell(row=2, column=col, value=cat["label"])
        cell.font = head_font
        cell.alignment = center
        cell.fill = group_fill
        col += span
    ws.cell(row=2, column=total_col, value="Total")
    ws.cell(row=2, column=grade_col, value="Grade")
    for cell in ws[2]:
        cell.fill = header_fill
        cell.border = border
        cell.font = head_font
        cell.alignment = center

    col = first_comp_col
    for comp in components:
        codes = ", ".join(comp.get("so_pi") or so_pi_labels(cpl_pis_of(comp)))
        c = ws.cell(row=3, column=col, value=f"{comp['name']} ({codes}) {comp['weight']}%")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.font = Font(size=9)
        c.border = border
        c.fill = header_fill
        col += 1
    ws.cell(row=3, column=total_col).border = border
    ws.cell(row=3, column=grade_col).border = border

    ws.cell(row=4, column=1, value="Bobot (%)")
    ws.cell(row=4, column=2, value="")
    ws.cell(row=4, column=3, value="")
    col = first_comp_col
    for comp in components:
        ws.cell(row=4, column=col, value=comp["weight"]).border = border
        col += 1
    ws.cell(row=4, column=total_col).border = border
    ws.cell(row=4, column=grade_col).border = border
    ws.cell(row=4, column=1).font = head_font

    data_start = 5
    for i, st in enumerate(students):
        r = data_start + i
        ws.cell(row=r, column=1, value=st["row_no"])
        ws.cell(row=r, column=2, value=st["nim"])
        ws.cell(row=r, column=3, value=st["name"])
        col = first_comp_col
        for comp in components:
            val = score_rows.get(st["id"], {}).get(comp["id"])
            ws.cell(row=r, column=col, value=val)
            col += 1
        first_letter = get_column_letter(first_comp_col)
        last_letter = get_column_letter(last_comp_col)
        total_letter = get_column_letter(total_col)
        if include_formulas and len(components) > 0:
            ws.cell(row=r, column=total_col, value=f"=ROUND(SUMPRODUCT({first_letter}{r}:{last_letter}{r},{first_letter}$4:{last_letter}$4)/100,2)")
            ws.cell(row=r, column=grade_col, value=f'={_grade_formula(total_letter + str(r))}')
        else:
            total = _compute_total(score_rows.get(st["id"], {}), components)
            ws.cell(row=r, column=total_col, value=round(total, 2))
            ws.cell(row=r, column=grade_col, value=grade_of_score(total))
        for cell in ws[r]:
            cell.border = border
            if cell.column > 1:
                cell.alignment = center

    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    for row in ws.iter_rows(min_row=data_start, max_row=data_start + len(students) - 1):
        for cell in row:
            if cell.column in (total_col, grade_col):
                cell.protection = Protection(locked=True)
            else:
                cell.protection = Protection(locked=False)
        ws.protection.enable()

    ws.freeze_panes = "D5"
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 26
    for ci in range(first_comp_col, last_comp_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.column_dimensions[get_column_letter(total_col)].width = 10
    ws.column_dimensions[get_column_letter(grade_col)].width = 8
    return wb


def workbook_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def load_workbook_bytes(data):
    return load_workbook(io.BytesIO(data), data_only=True)


# Category key -> SID column header (order matters)
SID_CATEGORY_COLUMNS = [
    ("partisipatif", "PARTISIPATIF"),
    ("tugas", "TUGAS"),
    ("proyek", "PROYEK"),
    ("kuis", "QUIZ"),
    ("ats", "UTS"),
    ("aas", "UAS"),
]


def _category_score(scores_map, components, category_id):
    """Weighted average (0-100) of a category's scored components; unfilled = 0."""
    s = 0.0
    w = 0.0
    for comp in components:
        if comp["category_id"] != category_id:
            continue
        val = scores_map.get(comp["id"])
        if val is not None:
            s += val * comp["weight"]
            w += comp["weight"]
    return (s / w) if w else 0.0


def build_sid_workbook(data):
    """Simplified Rubrik SID export:
    columns NO NIM NAMA + one summarized column per level-1 category.
    Pure data + borders only, headers from row 1 (A1), no styling/protection."""
    components = data["components"]
    categories = data["categories"]
    students = data["students"]
    score_rows = {}
    for sc in data["scores"]:
        score_rows.setdefault(sc["student_id"], {})[sc["component_id"]] = sc["score"]

    cat_id_by_key = {c["key"]: c["id"] for c in categories}

    wb = Workbook()
    ws = wb.active
    ws.title = "Rubrik SID"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["NO", "NIM", "NAMA"] + [label for _, label in SID_CATEGORY_COLUMNS]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).border = border

    for i, st in enumerate(students):
        r = 2 + i
        ws.cell(row=r, column=1, value=st["row_no"])
        ws.cell(row=r, column=2, value=st["nim"])
        ws.cell(row=r, column=3, value=st["name"])
        col = 4
        for key, _label in SID_CATEGORY_COLUMNS:
            cat_id = cat_id_by_key.get(key)
            val = _category_score(score_rows.get(st["id"], {}), components, cat_id) if cat_id else 0.0
            ws.cell(row=r, column=col, value=round(val, 2))
            col += 1
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = border

    ws.freeze_panes = "A2"
    return wb
